import os
import re
import pandas as pd
import xlrd
import csv
from openpyxl import load_workbook

class OD_data_obj:
    def __init__(self, nrows, ncols,starting_row,starting_col):
        self.nrows = nrows
        self.ncols = ncols
        self.index = {}
        #Index is a dictionary that links well labels (eg. A1, G3) to list indices of self.data
        alphabet = ['A','B','C','D','E','F','G','H']
        for jj in range(nrows):
            row_label = alphabet[alphabet.index(starting_row)+jj]
            for kk in range(ncols):
                well_label = row_label+str(kk+starting_col) 
                self.index[well_label] = jj*ncols+kk                 
        self.data = []  
        #This is a list of lists that represent the time series for each well. 

class well_array: 
    def __init__(self,wells):
        self.wells = wells
        self.nrows = len(self.wells)
        self.ncols = len(self.wells[0])


def get_OD_data(dirname,fname, ntimes, nrows, ncols,starting_row,starting_col):
    fname_full = os.path.normpath(dirname + fname)
    lines = open(fname_full,'r').readlines()
    
    #Grab time points and turn into list
    time_list = lines[1]
    time_list = time_list.split(',')[0:ntimes]
    #removes units and converts to minutes
    if time_list[0][-1]=='"': 
        #when conversion not done in excel, double quotes are around the entry
        time_list = [int(time_val.strip('"')[0:-1])/60.0  for time_val in time_list]
    else: 
        time_list = [int(time_val[0:-1])/60.0  for time_val in time_list]
    #Import data as 8 rows of 12.  
    #Set data index to first row in which data can appear. 
    nn = 3 
    
    #make OD_data_obj object to store data.
    OD_data = OD_data_obj(nrows, ncols, starting_row, starting_col)
    
    if (nrows == 8) and (ncols == 12):
        for jj in range(nrows):
            for kk in range(ncols):
                data_line = lines[nn]
                data_line = re.split("[,\n]",data_line)
                if data_line[0][0] == '"':
                    OD_data.data.append([float(data_val.strip('"')) for data_val in data_line[0:ntimes]])
                else: 
                    OD_data.data.append([float(data_val) for data_val in data_line[0:ntimes]])
                nn +=1

    else:  
        #this is tailored for data sets that come from plates in which outside wells are not measured. 
        first_val = lines[nn].split(',')[0]
        while first_val == '':
            nn += 1
            first_val = lines[nn].split(',')[0]
        
        #This assumes two new lines between each new data point.
        for jj in range(nrows):
            for kk in range(ncols):
                data_line = lines[nn]
                data_line = re.split("[,\n]",data_line)
                if data_line[0][0] == '"':
                    OD_data.data.append([float(data_val.strip('"')) for data_val in data_line[0:ntimes]])
                else: 
                    OD_data.data.append([float(data_val) for data_val in data_line[0:ntimes]])
                nn +=1
            nn +=2
    
    return OD_data, time_list
    
              
def all_wells(letter_range,number_range):
    wells = []    
    for jj in range(len(letter_range)):
        well_cols = []
        for kk in range(len(number_range)):
            well_cols.append(letter_range[jj]+str(number_range[kk]))
        wells.append(well_cols)
    return wells
    
def load_spark10_data(fname):
    wb = load_workbook(filename = fname)
    #Cycle through cells until I find 'Cycle Nr.'
    #Assumes first worksheet is the correct worksheet. 
    data_sheet = wb[wb.get_sheet_names()[0]]
    first_column = data_sheet.columns[0]
    for jj in range(len(first_column)): 
        cell = first_column[jj]
        if cell.value == 'Cycle Nr.':
            break
    
    #Record location 
    data_table_first_row = cell.row
    
    #Find name of last column that contains data
    header_row = data_sheet.rows[data_table_first_row-1]
    data_table_last_column = header_row[-1].column
    
    #Find row of last timepoint
    for jj in range(data_table_first_row, len(first_column)): 
        cell = first_column[jj]
        if cell.value == None:
            break
    
    data_table_last_row = cell.row -1
            
    #Load Relevant table
    data_table = tuple(data_sheet[('A'+str(data_table_first_row)):(data_table_last_column + str(data_table_last_row))])
    
    #Convert to dataframe
    header = [cell.value for cell in data_table[0]]
    data_table_values = []
    for row in data_table[1:]:
        data_table_values_row = []
        for cell in row:
            #print cell.value
            data_table_values_row.append(cell.value)
        data_table_values.append(data_table_values_row)
    
    data = pd.DataFrame(data_table_values, columns = header)
    data.set_index('Cycle Nr.',inplace = True)   
    
    return data
    
    
def load_thermo_data(fname):

    wb = load_workbook(filename = fname)
    #Cycle through cells until I find 'Cycle Nr.'
    data_sheet = wb[wb.get_sheet_names()[0]]
    first_column = data_sheet.columns[0]
    for jj in range(len(first_column)): 
        cell = first_column[jj]
        if cell.value == 'Reading':
            break
    
    #Record location 
    data_table_first_row = cell.row
    
    #Find name of last column that contains data
    header_row = data_sheet.rows[data_table_first_row-1]
    data_table_last_column = header_row[-2].column  #The thermosci data picks up a blank column, so we use -2 instead of -1. 
    
    #Find row of last timepoint
    for jj in range(data_table_first_row, len(first_column)): 
        cell = first_column[jj]
        if cell.value == None:
            break
    
    data_table_last_row = cell.row -2
    
    #Load Relevant table
    data_table = tuple(data_sheet[('A'+str(data_table_first_row)):(data_table_last_column + str(data_table_last_row))])
    
    #Convert to dataframe
    header = [cell.value for cell in data_table[0]]
    #Convert header to match spark10 output 
    header[1] = 'Time [s]'
    alphabet = ['A','B','C','D','E','F','G','H']
    ncols = 12
    for jj in range(len(alphabet)):
        row_label = alphabet[jj]
        for kk in range(1,ncols+1):
            header[1+jj*ncols+kk]= row_label+str(kk) 
            
    
    data_table_values = []
    for row in data_table[1:]:
        data_table_values_row = []
        for cell in row:
            #print cell.value
            data_table_values_row.append(cell.value)
        data_table_values.append(data_table_values_row)
    
    data = pd.DataFrame(data_table_values, columns = header)
    data.set_index('Reading',inplace = True)   

    return data    

def xlsx_to_csv(fname):
    #converts xlsx files from the lim lab plate reader to .csv files.  Cannot convert degree symbol correctly. 
    x =  xlrd.open_workbook(os.path.normpath(fname))
    fname_out = fname.split('.')[0]+".csv"
    x1 = x.sheet_by_name('Magellan Sheet 1')
    csvfile = open(os.path.normpath(fname_out),'wb')
    writecsv = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
    for rownum in xrange(x1.nrows):
        row = x1.row_values(rownum)
        row_encoded = []
        for s in row: 
            if isinstance(s,unicode):
                row_encoded.append(s.encode('utf-8'))
            else: 
                row_encoded.append(s)
        writecsv.writerow(row_encoded)
    csvfile.close()
    return fname_out